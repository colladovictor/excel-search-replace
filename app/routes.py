from flask import Blueprint, jsonify, render_template, request
import os
from openpyxl import load_workbook

main = Blueprint('main', __name__)

@main.route('/')
def index():
    return render_template('index.html')

def search_excel_files(directory, search_text, replace_text=None):
    results = []
    for root, dirs, files in os.walk(directory):
        for file in files:
            if file.endswith('.xlsx'):
                filepath = os.path.join(root, file)
                try:
                    workbook = load_workbook(filename=filepath, data_only=True)
                    for sheet_name in workbook.sheetnames:
                        sheet = workbook[sheet_name]
                        for rowIndex, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                            if any(search_text.lower() in (str(cell).lower() if cell else '') for cell in row):
                                row_content = [str(cell) if cell else '' for cell in row]
                                if replace_text is not None:
                                    for cell in sheet[rowIndex]:
                                        if cell.value and search_text.lower() in str(cell.value).lower():
                                            cell.value = str(cell.value).replace(search_text, replace_text)
                                    workbook.save(filepath)
                                results.append({
                                    'file': os.path.relpath(filepath, directory),
                                    'sheet': sheet_name,
                                    'row': rowIndex,
                                    'content': ' | '.join(row_content)
                                })
                except Exception as e:
                    print(f"Error processing {filepath}: {e}")
    return results

@main.route('/search', methods=['POST'])
def search():
    data = request.get_json()
    directory = data['directoryPath']
    search_text = data['searchText']
    results = search_excel_files(directory, search_text)
    return jsonify(results)

@main.route('/replace', methods=['POST'])
def replace():
    data = request.get_json()
    directory = data['directoryPath']
    search_text = data['searchText']
    replace_text = data['replaceText']
    results = search_excel_files(directory, search_text, replace_text)
    return jsonify(results)
